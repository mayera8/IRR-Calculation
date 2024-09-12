import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

#pd.set_option('display.max_columns', None)

#Path to excel file
example_model_path = r''


#Read excel file specifically pulling cashflows and dates
example_model = pd.read_excel(example_model_path, sheet_name='PE Returns analysis', index_col= 0, usecols= 'A, C:H', skiprows= range(0, 13), nrows=7)
#print(example_model)

total_cashflows = example_model.iloc[5]
#print(total_cashflows)
#print(total_cashflows.dtypes)


#Dataframe of cashflows and dates, clean up to create list of dates and cashflows
total_cashflows = pd.DataFrame(total_cashflows)
#print(total_cashflows)
#print(total_cashflows.dtypes)

total_cashflows = total_cashflows.reset_index()
total_cashflows = total_cashflows.rename(columns={'index':'Date'})
#print(total_cashflows)
#print(total_cashflows.dtypes)


#Create list for Total Cashflows and Dates
total_cashflows_list = total_cashflows['Total Cashflows'].tolist()
#print(total_cashflows_list)

date_list = total_cashflows['Date'].tolist()
#print(date_list)


#Convert dates into time intervals in terms of years since first cash flow
start_date = date_list[0]
times = [(x - start_date).days /365.0 for x in date_list]
#print(times)


#XIRR Calculation - for irregular cash flows
def xirr_calc(total_cashflows_list, times, initial_guess=0.1):
	rate = initial_guess
	tolerance = 1e-6
	max_iterations = 1000
	iteration = 0

	while iteration < max_iterations:
		npv = sum(cf / (1 + rate) ** t for cf, t in zip(total_cashflows_list, times))
		
		if abs(npv) < tolerance:
			return rate

		rate += npv / 1000 if npv > 0 else npv / 10000
		iteration += 1

	return rate

xirr = xirr_calc(total_cashflows_list, times) * 100
print(xirr)


#IRR Calculation - regular cash flows
def irr_calc(total_cashflows_list, initial_guess=0.1):
	rate = initial_guess
	tolerance = 1e-6
	max_iterations = 1000
	iteration = 0

	while iteration < max_iterations:
		npv = sum(cf / (1 + rate) ** i for i, cf in enumerate(total_cashflows_list))

		if abs(npv) < tolerance:
			return rate

		rate += npv / 1000 if npv > 0 else npv / 10000
		iteration += 1

	return rate

#irr = irr_calc(total_cashflows_list) * 100
#print(irr)


#Save to Excel file without changing formatting
load = load_workbook(example_model_path)
pe_returns_analysis = load['PE Returns analysis']

pe_returns_analysis.cell(row= 23, column= 3, value= xirr)

load.save(example_model_path)